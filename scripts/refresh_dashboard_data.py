from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import statistics
import subprocess
import tempfile
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
BUNDLE_DIR = SCRIPT_DIR.parent
DEFAULT_WORKBOOK = (BUNDLE_DIR.parent / "03_Deliverables" / "Chipunza_Project_Control_Live.xlsx").resolve()
DEFAULT_OUTPUT = (BUNDLE_DIR / "dashboard_data.json").resolve()
TIMEZONE = dt.timezone(dt.timedelta(hours=2), name="SAST")
CREATE_NO_WINDOW = getattr(subprocess, "CREATE_NO_WINDOW", 0)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build dashboard JSON from the Chipunza control workbook.")
    parser.add_argument("--workbook", help="Local workbook path.")
    parser.add_argument("--workbook-url", help="Optional public workbook URL.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output JSON path.")
    return parser.parse_args()


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def public_text(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return text
    text = re.sub(r"\s*(?:\+|,|/|&)?\s*Codex\s*(?:\+|,|/|&)?\s*", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s{2,}", " ", text).strip(" ,/&+")
    return text


def parse_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value).replace(",", "").replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: Any) -> dt.date | None:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = clean_text(value)
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def iso_date(value: dt.date | None) -> str | None:
    return value.isoformat() if value else None


def fmt_date_label(value: str | None) -> str:
    if not value:
        return "-"
    try:
        return dt.date.fromisoformat(value).strftime("%d %b %Y")
    except ValueError:
        return value


def tone_for_status(status: str) -> str:
    status = clean_text(status)
    if status == "Complete":
        return "good"
    if status in {"In Progress"}:
        return "info"
    if status in {"Blocked", "At Risk"}:
        return "bad"
    return "warn"


def workbook_rows(ws) -> list[dict[str, Any]]:
    headers = [clean_text(cell.value) for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        row = {headers[i]: values[i] for i in range(len(headers))}
        rows.append(row)
    return rows


def load_workbook_from_source(workbook: str | None, workbook_url: str | None) -> tuple[Path, Any]:
    if workbook:
        path = Path(workbook).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")
        snapshot = create_snapshot(path)
        return path, load_workbook(snapshot, data_only=False)

    if DEFAULT_WORKBOOK.exists():
        snapshot = create_snapshot(DEFAULT_WORKBOOK)
        return DEFAULT_WORKBOOK, load_workbook(snapshot, data_only=False)

    if workbook_url:
        with urllib.request.urlopen(workbook_url) as response:  # noqa: S310
            data = response.read()
        temp_dir = Path(tempfile.mkdtemp(prefix="chipunza_dashboard_"))
        temp_path = temp_dir / "Chipunza_Project_Control_Live.xlsx"
        temp_path.write_bytes(data)
        return temp_path, load_workbook(temp_path, data_only=False)

    raise FileNotFoundError("No workbook found. Provide --workbook or store the workbook in ../03_Deliverables.")


def create_snapshot(source_path: Path) -> Path:
    if os.name != "nt":
        return source_path

    helper = SCRIPT_DIR / "save_excel_snapshot.ps1"
    temp_dir = Path(tempfile.mkdtemp(prefix="chipunza_workbook_"))
    target = temp_dir / source_path.name
    command = [
        "powershell",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(helper),
        "-SourcePath",
        str(source_path),
        "-TargetPath",
        str(target),
    ]
    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    startupinfo.wShowWindow = 0
    result = subprocess.run(command, text=True, capture_output=True, creationflags=CREATE_NO_WINDOW, startupinfo=startupinfo)
    if result.returncode == 0 and target.exists():
        return target
    return source_path


def build_summary(tasks: list[dict[str, Any]], decisions: list[dict[str, Any]], milestones: list[dict[str, Any]], roles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    progress_values = [parse_float(row.get("Percent_Complete")) for row in tasks]
    progress_values = [value for value in progress_values if value is not None]
    avg_progress = statistics.fmean(progress_values) if progress_values else 0.0
    blocked_count = sum(1 for row in tasks if clean_text(row.get("Status")) == "Blocked")
    in_progress_count = sum(1 for row in tasks if clean_text(row.get("Status")) == "In Progress")
    open_decisions = sum(1 for row in decisions if clean_text(row.get("Status")) == "Open")
    next_milestone = None
    dated = []
    today = dt.date.today()
    for row in milestones:
        mdate = parse_date(row.get("date"))
        if mdate and clean_text(row.get("status")) != "Complete" and mdate >= today:
            dated.append((mdate, row))
    if dated:
        dated.sort(key=lambda item: item[0])
        next_milestone = dated[0][1]

    site_start = next((m for m in milestones if clean_text(m.get("label")) == "Formal site start"), None)
    handover = next((m for m in milestones if clean_text(m.get("label")) == "Owner handover"), None)
    onsite_lead = next((r for r in roles if "On-Site" in clean_text(r.get("Role")) or "Dherifah" in clean_text(r.get("Name"))), None)

    return [
        {"label": "Site Start", "value": fmt_date_label(site_start.get("date")) if site_start else "01-May-2026", "detail": "Formal site mobilisation", "tone": "info"},
        {"label": "Planned Finish", "value": fmt_date_label(handover.get("date")) if handover else "-", "detail": "Current programme target", "tone": "warn"},
        {"label": "Overall Progress", "value": f"{avg_progress:.0%}", "detail": "Average task completion", "tone": "good" if avg_progress >= 0.75 else "warn" if avg_progress >= 0.35 else "bad"},
        {"label": "In Progress", "value": str(in_progress_count), "detail": "Active work items", "tone": "info"},
        {"label": "Blocked Tasks", "value": str(blocked_count), "detail": "Immediate constraints", "tone": "bad" if blocked_count else "good"},
        {"label": "Open Decisions", "value": str(open_decisions), "detail": "Still awaiting closure", "tone": "bad" if open_decisions > 3 else "warn" if open_decisions else "good"},
        {"label": "On-Site Lead", "value": public_text(onsite_lead.get("Name")) if onsite_lead else "Del", "detail": "Day-to-day site coordination", "tone": "info"},
        {"label": "Next Milestone", "value": public_text(next_milestone.get("label")) if next_milestone else "-", "detail": fmt_date_label(next_milestone.get("date")) if next_milestone else "No future milestone set", "tone": "info"},
    ]


def build_phase_progress(tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    phases: dict[str, list[dict[str, Any]]] = {}
    for row in tasks:
        phases.setdefault(clean_text(row.get("Phase")), []).append(row)

    results = []
    for phase, rows in phases.items():
        progress = [parse_float(row.get("Percent_Complete")) for row in rows]
        progress = [value for value in progress if value is not None]
        completed = sum(1 for row in rows if clean_text(row.get("Status")) == "Complete")
        blocked = sum(1 for row in rows if clean_text(row.get("Status")) == "Blocked")
        average = statistics.fmean(progress) if progress else 0.0
        results.append(
            {
                "phase": public_text(phase),
                "progress": average,
                "completed": completed,
                "total": len(rows),
                "blocked": blocked,
                "tone": "good" if average >= 0.8 else "warn" if average >= 0.35 else "bad",
            }
        )
    return results


def build_upcoming_tasks(tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sortable = []
    for row in tasks:
        start = parse_date(row.get("Start_Date"))
        if start:
            sortable.append((start, row))
    sortable.sort(key=lambda item: item[0])
    results = []
    for _, row in sortable[:12]:
        results.append(
            {
                "id": clean_text(row.get("Task_ID")),
                "phase": public_text(row.get("Phase")),
                "task": public_text(row.get("Task")),
                "owner": public_text(row.get("Owner")),
                "start": iso_date(parse_date(row.get("Start_Date"))),
                "finish": iso_date(parse_date(row.get("Finish_Date"))),
                "status": public_text(row.get("Status")),
                "progress": parse_float(row.get("Percent_Complete")) or 0.0,
                "critical": clean_text(row.get("Critical")) == "Yes",
                "notes": public_text(row.get("Notes")),
            }
        )
    return results


def build_blocked_tasks(tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    blocked = []
    for row in tasks:
        if clean_text(row.get("Status")) in {"Blocked", "At Risk"}:
            blocked.append(
                {
                    "id": clean_text(row.get("Task_ID")),
                    "task": public_text(row.get("Task")),
                    "owner": public_text(row.get("Owner")),
                    "status": public_text(row.get("Status")),
                    "dependency": public_text(row.get("Dependency")),
                    "notes": public_text(row.get("Notes")),
                }
            )
    return blocked


def build_gantt(tasks: list[dict[str, Any]]) -> dict[str, Any]:
    dated = []
    for row in tasks:
        start = parse_date(row.get("Start_Date"))
        finish = parse_date(row.get("Finish_Date"))
        if start and finish:
            dated.append((start, finish, row))
    if not dated:
        return {"startDate": None, "endDate": None, "tasks": []}

    start_date = min(item[0] for item in dated)
    end_date = max(item[1] for item in dated)
    gantt_rows = []
    for index, (start, finish, row) in enumerate(sorted(dated, key=lambda item: (item[0], item[1]))):
        gantt_rows.append(
            {
                "row": index + 1,
                "id": clean_text(row.get("Task_ID")),
                "task": public_text(row.get("Task")),
                "phase": public_text(row.get("Phase")),
                "owner": public_text(row.get("Owner")),
                "start": iso_date(start),
                "finish": iso_date(finish),
                "status": public_text(row.get("Status")),
                "progress": parse_float(row.get("Percent_Complete")) or 0.0,
                "critical": clean_text(row.get("Critical")) == "Yes",
            }
        )
    return {"startDate": iso_date(start_date), "endDate": iso_date(end_date), "tasks": gantt_rows}


def build_milestone_payload(milestones: list[dict[str, Any]], tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    task_lookup = {clean_text(row.get("Task_ID")): row for row in tasks}
    results = []
    for row in milestones:
        task_id = clean_text(row.get("Task_ID"))
        source = task_lookup.get(task_id, {})
        target_date = parse_date(source.get("Finish_Date"))
        if clean_text(row.get("Milestone")) == "Formal site start":
            target_date = parse_date(source.get("Start_Date")) or target_date
        results.append(
            {
                "label": public_text(row.get("Milestone")),
                "date": iso_date(target_date),
                "status": public_text(source.get("Status")),
                "owner": public_text(source.get("Owner")),
                "notes": public_text(row.get("Notes")),
            }
        )
    return results


def build_payload(workbook_path: Path, workbook) -> dict[str, Any]:
    tasks = workbook_rows(workbook["Task_Plan"])
    budgets = workbook_rows(workbook["Budget_Register"])
    decisions = workbook_rows(workbook["Decision_Log"])
    milestone_rows = workbook_rows(workbook["Milestones"])
    roles = workbook_rows(workbook["Team_Roles"])
    milestones = build_milestone_payload(milestone_rows, tasks)

    source_modified = dt.datetime.fromtimestamp(workbook_path.stat().st_mtime, tz=TIMEZONE).isoformat() if workbook_path.exists() else None
    generated = source_modified

    return {
        "title": "Chipunza House Project Dashboard",
        "subtitle": "Public contractor view for programme, milestones, constraints, and site coordination",
        "sourceName": workbook_path.name,
        "generatedAt": generated,
        "sourceModifiedAt": source_modified,
        "refreshSeconds": 60,
        "summaryCards": build_summary(tasks, decisions, milestones, roles),
        "phaseProgress": build_phase_progress(tasks),
        "milestones": milestones,
        "upcomingTasks": build_upcoming_tasks(tasks),
        "blockedTasks": build_blocked_tasks(tasks),
        "gantt": build_gantt(tasks),
        "budget": {
            "items": [
                {
                    "package": public_text(row.get("Package")),
                    "status": public_text(row.get("Status")),
                    "owner": public_text(row.get("Owner")),
                    "notes": public_text(row.get("Notes")),
                }
                for row in budgets
            ]
        },
        "decisions": [
            {
                "id": clean_text(row.get("Decision_ID")),
                "decision": public_text(row.get("Decision")),
                "requiredBy": iso_date(parse_date(row.get("Required_By"))),
                "owner": public_text(row.get("Owner")),
                "status": public_text(row.get("Status")),
                "impact": public_text(row.get("Impact")),
            }
            for row in decisions
        ],
        "roles": [
            {
                "name": public_text(row.get("Name")),
                "role": public_text(row.get("Role")),
                "location": public_text(row.get("Location")),
                "responsibility": public_text(row.get("Primary Responsibility")),
            }
            for row in roles
        ],
    }


def refresh_dashboard_data(workbook: str | None = None, workbook_url: str | None = None, output: str | Path = DEFAULT_OUTPUT) -> Path:
    workbook_path, loaded = load_workbook_from_source(workbook, workbook_url)
    payload = build_payload(workbook_path, loaded)
    output_path = Path(output).expanduser().resolve()
    output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return output_path


def main() -> None:
    args = parse_args()
    refresh_dashboard_data(workbook=args.workbook, workbook_url=args.workbook_url, output=args.output)


if __name__ == "__main__":
    main()
